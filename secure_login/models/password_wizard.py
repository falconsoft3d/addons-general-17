# coding: utf-8
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    This module copyright (C) 2018 Marlon Falcón Hernandez
#    (<http://www.falconsolutions.cl>).
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################

import logging
import random
import string
import base64
import io
from datetime import datetime

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

_SPECIAL_CHARS = '!@#$%*'
_PASSWORD_LENGTH = 12


def _generate_secure_password():
    """Genera una contraseña segura usando SystemRandom (criptográficamente seguro).
    Garantiza al menos: 1 mayúscula, 1 minúscula, 1 dígito, 1 carácter especial.
    """
    rng = random.SystemRandom()
    all_chars = string.ascii_letters + string.digits + _SPECIAL_CHARS
    while True:
        pwd = [rng.choice(all_chars) for _ in range(_PASSWORD_LENGTH)]
        if (
            any(c.isupper() for c in pwd)
            and any(c.islower() for c in pwd)
            and any(c.isdigit() for c in pwd)
            and any(c in _SPECIAL_CHARS for c in pwd)
        ):
            return ''.join(pwd)


class SecurePasswordWizard(models.TransientModel):
    _name = 'secure.password.wizard'
    _description = u'Asignacion Masiva de Contrasenas Seguras'

    user_ids = fields.Many2many(
        'res.users',
        string='Usuarios',
        domain=[('active', '=', True), ('share', '=', False)],
    )
    send_email = fields.Boolean(
        string=u'Enviar contrasena por correo a cada usuario',
        default=False,
    )

    def action_generate_passwords(self):
        self.ensure_one()

        if not self.user_ids:
            raise UserError(_('Debe seleccionar al menos un usuario.'))

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError(_(
                'La librería openpyxl no está instalada en el servidor.\n'
                'Ejecute: pip install openpyxl'
            ))

        # Generar contraseñas y actualizar usuarios
        # tupla: (nombre, login, pwd, email_usuario)
        rows = []
        for user in self.user_ids:
            pwd = _generate_secure_password()
            user.sudo().write({'password': pwd})
            rows.append((user.name or '', user.login or '', pwd, user.email or ''))

        _logger.info(
            'secure_login: contraseñas actualizadas masivamente para %d usuario(s) '
            'por el usuario %s (id=%d)',
            len(rows), self.env.user.login, self.env.uid,
        )

        # Enviar correo si el check está activo
        if self.send_email:
            for name, login, pwd, email in rows:
                if not email:
                    _logger.warning(
                        'secure_login: el usuario %s no tiene email configurado, '
                        'se omite el envio de correo', login,
                    )
                    continue
                body_html = (
                    u'<p>Estimado/a <strong>%s</strong>,</p>'
                    u'<p>Su contrasena de acceso al sistema ha sido actualizada '
                    u'por el administrador.</p>'
                    u'<table style="border-collapse:collapse;margin:12px 0;">'
                    u'<tr><td style="padding:4px 12px 4px 0;"><strong>Usuario:</strong></td>'
                    u'<td style="padding:4px 0;">%s</td></tr>'
                    u'<tr><td style="padding:4px 12px 4px 0;"><strong>Nueva contrasena:</strong></td>'
                    u'<td style="padding:4px 0;font-family:monospace;font-size:14px;">%s</td></tr>'
                    u'</table>'
                    u'<p style="color:#888;font-size:12px;">Por seguridad, le recomendamos '
                    u'cambiar esta contrasena en su proximo inicio de sesion.</p>'
                ) % (name, login, pwd)
                try:
                    mail = self.env['mail.mail'].create({
                        'subject': u'Nueva contrasena de acceso al sistema',
                        'body_html': body_html,
                        'email_to': email,
                        'auto_delete': True,
                    })
                    mail.send()
                except Exception:
                    _logger.exception(
                        'secure_login: error al enviar correo a %s', email,
                    )

        # Construir el archivo XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Contrasenas'

        thin_side = Side(style='thin', color='CCCCCC')
        border = Border(
            left=thin_side, right=thin_side,
            top=thin_side, bottom=thin_side,
        )
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='2E75B6')
        header_align = Alignment(horizontal='center', vertical='center')

        headers = ['Nombre', 'Login / Correo', 'Contrasena']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        alt_fill = PatternFill('solid', fgColor='DCE6F1')
        data_align = Alignment(horizontal='left', vertical='center')
        for row_idx, (name, login, pwd, _email) in enumerate(rows, 2):
            row_fill = alt_fill if row_idx % 2 == 0 else None
            for col, value in enumerate([name, login, pwd], 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = data_align
                if row_fill:
                    cell.fill = row_fill

        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 36
        ws.column_dimensions['C'].width = 18
        ws.freeze_panes = 'A2'

        # Serializar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read())
        file_name = 'contrasenas_%s.xlsx' % datetime.now().strftime('%Y%m%d_%H%M%S')

        # Crear adjunto para descarga segura
        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'datas': file_data,
            'datas_fname': file_name,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % attachment.id,
            'target': 'new',
        }
